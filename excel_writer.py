from openpyxl import Workbook
import os

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "data", "Podcast_Scraper_Database.xlsx")


def ensure_excel_exists():
    """Create Excel file with headers if it doesn't exist"""
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    
    if not os.path.exists(EXCEL_PATH):
        create_excel_template()


def create_excel_template():
    """Create a fresh Excel file with headers"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Episode Database"
    
    # Headers
    headers = [
        "Episode #", "Channel Name", "Title", "Guest Name", "Guest Role",
        "Guest Company", "Industry", "Description", "Publish Date", "Duration (min)",
        "Views", "Likes", "Comments", "Thumbnail URL", "Title Word Count",
        "Video URL", "Tags", "Engagement Rate %"
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = header
    
    wb.save(EXCEL_PATH)


def write_to_excel(episodes, channel_name):
    """
    Write episodes to Excel file.
    OVERWRITES previous data to prevent mixing different channels.
    """
    
    # Ensure directory exists
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    
    # Create fresh workbook (overwrites old file)
    wb = Workbook()
    ws = wb.active
    ws.title = "Episode Database"
    
    # Write headers
    headers = [
        "Episode #", "Channel Name", "Title", "Guest Name", "Guest Role",
        "Guest Company", "Industry", "Description", "Publish Date", "Duration (min)",
        "Views", "Likes", "Comments", "Thumbnail URL", "Title Word Count",
        "Video URL", "Tags", "Engagement Rate %"
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = header
    
    # Write episode data
    next_row = 2
    
    for episode in episodes:
        ws.cell(row=next_row, column=1).value = episode["episode_number"]
        ws.cell(row=next_row, column=2).value = channel_name
        ws.cell(row=next_row, column=3).value = episode["title"]
        
        ws.cell(row=next_row, column=4).value = episode["guest_name"]
        ws.cell(row=next_row, column=5).value = episode["guest_role"]
        ws.cell(row=next_row, column=6).value = episode["guest_company"]
        ws.cell(row=next_row, column=7).value = episode["industry"]
        
        ws.cell(row=next_row, column=8).value = episode["description"]
        
        ws.cell(row=next_row, column=9).value = episode["publish_date"]
        ws.cell(row=next_row, column=10).value = episode["duration"]
        
        ws.cell(row=next_row, column=11).value = episode["views"]
        ws.cell(row=next_row, column=12).value = episode["likes"]
        ws.cell(row=next_row, column=13).value = episode["comments"]
        
        ws.cell(row=next_row, column=14).value = episode["thumbnail"]
        ws.cell(row=next_row, column=15).value = episode["title_word_count"]
        
        ws.cell(row=next_row, column=16).value = episode["video_url"]
        ws.cell(row=next_row, column=17).value = episode["tags"]
        
        ws.cell(row=next_row, column=18).value = episode["engagement_rate"]
        
        next_row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                cell_length = len(str(cell.value or ""))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save
    wb.save(EXCEL_PATH)